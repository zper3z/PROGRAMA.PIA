import requests
import openpyxl 
import os

def crear_carpetaListas(ruta = os.getcwd()):
    ruta = os.path.join(ruta, "Listas")
    if not os.path.exists(ruta):
        os.makedirs(ruta)

def importar_listas(ruta_raiz = os.getcwd()):
    carpeta = os.path.join(ruta_raiz, "Listas")
    listas = {}
    try:
        archivos = os.listdir(carpeta)
        archivos_xml =  [documento for documento in archivos if documento.endswith('.xlsx')]
        for documento in archivos_xml:
            ruta = os.path.join(carpeta,documento)
            documento = documento.replace(".xlsx","")
            listas[documento] = ruta
        return listas
    except FileNotFoundError:
        print("Error: El archivo no se encontró.")

    except Exception as e:
        print(f"Otro error: {e}")
	
def consultar_lista(ruta):
    try:
        documento = openpyxl.load_workbook(ruta)
        hoja = documento.active
        filas = hoja.iter_rows(values_only=True)
        lista = [list(fila) for fila in filas]
        documento.close()
        return lista
    except FileNotFoundError:
        print("No existe ese archivo")

def buscar(nombre_pokemon):
    if nombre_pokemon != "":
        url = f"https://pokeapi.co/api/v2/pokemon/{nombre_pokemon}"
        response = requests.get(url)
        informacion = []
            
        if response.status_code == 200 :
            datos = response.json()
            nombre = datos["name"]
            habilidades = []
            habilidad_oc= ""
            tipos = [tipo["type"]["name"] for tipo in datos["types"]]
            for habilidad in datos["abilities"]:
                if habilidad["is_hidden"]:
                    habilidad_oc = habilidad["ability"]["name"]
                else:
                    habilidades.append(habilidad["ability"]["name"])
            
            if habilidad_oc == "":
                habilidad_oc = "No tiene"
            informacion = [nombre, [], [], habilidad_oc]
            informacion[1].extend(tipos)
            informacion[2].extend(habilidades)
                

            return informacion
        else:
            print(f"No se encontró ningún Pokémon con el nombre o número '{nombre_pokemon}'.")
            nombre = input("Ingresa el nombre o número del Pokémon: ")
            return buscar(nombre)
    else:
        print(f"Favor de ingresar los datos pedidos")
        nombre = input("Ingresa el nombre o número del Pokémon: ")
        return buscar(nombre)

def imprimir(info):
	nombre = info[0]
	tipos = info[1]
	habilidades = info[2]
	habilidad_oc = info[3]
	print(f"Nombre: {nombre}")
	print("Tipos:", ", ".join(tipos))
	print("Habilidades:", ", ".join(habilidades))
	print("Habilidad Oculta:", habilidad_oc)

def excel(info, lista, ruta = os.getcwd()):
    ruta = os.path.join(ruta, "Listas")
    try:
        libro_trabajo = openpyxl.Workbook()
        hoja = libro_trabajo.active
        
        hoja['A1'] = "Nombre"
        hoja['B1'] = "Tipos"
        hoja['C1'] = "Habilidades"
        hoja['D1'] = "Habilidad Oculta"

        for pokemon in info:
            datos = buscar(pokemon)
            hoja.append([datos[0], ", ".join(datos[1]), ", ".join(datos[2]),datos[3]])
        libro_trabajo.save(f"{ruta}\\{lista}.xlsx")
        libro_trabajo.close()
        print(f"La información de tu lista {lista} se ha guardado en {lista}.xlsx")
    except FileExistsError:
        print("Error: El excel ya existe.")

def ciclo_si_no(texto):
    opcion = input(f"{texto} (Si/No): ").lower()
    if opcion == "si":
        return True
    elif opcion == "no":
        return False
    else:
        return ciclo_si_no(texto)
    
def validador_opciones(opcion_final):
    opcion_final = int(opcion_final)
    opcion = input(f"Selecciona su opción (1-{opcion_final}): ")
    for num in range(1,opcion_final+1,1):
        if opcion == str(num):break
    else:
        print(f"Opción inválida")
        return validador_opciones(opcion_final)
    return int(opcion)
